import json
import sys
from openpyxl import load_workbook


def main() -> int:
    if len(sys.argv) < 2:
        print("Path to xlsx file is required", file=sys.stderr)
        return 1

    path = sys.argv[1]

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)

    if not header_row:
        print(json.dumps({"headers": [], "rows": []}, ensure_ascii=False))
        return 0

    headers = ["" if value is None else str(value) for value in header_row]
    rows = []

    for row in rows_iter:
        rows.append([
            "" if value is None else str(value)
            for value in row
        ])

    print(json.dumps({"headers": headers, "rows": rows}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
